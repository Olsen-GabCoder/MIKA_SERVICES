#!/usr/bin/env python3
"""
Import incrémental des lignes **matériau** des feuilles du barème Tout_Corps (.xls),
hors feuille 0 (coefficients d’éloignement).

- Même grille Excel que `merge_gros_oeuvre_missing.py` (colonnes N°, Matériaux, U, P.TTC, …).
- **Un corps d’état par feuille**, avec un `code` aligné sur le backend Kotlin
  (`BaremeImportService.getOrCreateCorpsEtat` : nom en majuscules, caractères non [A-Z0-9_] → _).
- Par défaut : **dédoublonnage uniquement en base** pour le corps ciblé
  (libellé + unité + fournisseur résolu). N’efface rien.
- Résolution des noms fournisseurs : liste des libellés du fichier **BASE DE DONNE F.xlsx**
  (comme le merge Gros-Œuvre), sinon nom tel quel.

Étapes recommandées : une feuille à la fois (`--sheet-index`), dry-run puis `--execute`.

Usage :
  python scripts/merge_tout_corps_sheets.py --sheet-index 2
  python scripts/merge_tout_corps_sheets.py --sheet-index 2 --execute
  python scripts/merge_tout_corps_sheets.py --from-index 2 --to-index 15 --execute

Prérequis : pymysql, openpyxl, xlrd ; `docs/BASE DE DONNE F.xlsx` + fichier `.xls` Tout_Corps ;
variables DB dans `backend/.env`.
"""
from __future__ import annotations

import argparse
import os
import re
import sys
import unicodedata
from decimal import Decimal

import openpyxl
import xlrd

ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
DOCS = os.path.join(ROOT, "docs")
ENV_PATH = os.path.join(ROOT, "backend", ".env")


def load_env(path: str) -> dict:
    out = {}
    if not os.path.isfile(path):
        return out
    with open(path, encoding="utf-8", errors="replace") as f:
        for line in f:
            line = line.strip()
            if not line or line.startswith("#") or "=" not in line:
                continue
            k, _, v = line.partition("=")
            k, v = k.strip(), v.strip().strip('"').strip("'")
            out[k] = v
    return out


def parse_jdbc_url(url: str) -> tuple[str, int, str]:
    import re as _re

    m = _re.match(r"jdbc:mysql://([^:/]+):(\d+)/([^?]+)", url)
    if not m:
        raise ValueError(f"URL JDBC MySQL non reconnue: {url[:60]}...")
    return m.group(1), int(m.group(2)), m.group(3)


def strip_accents(s: str) -> str:
    if not s:
        return ""
    return "".join(
        c for c in unicodedata.normalize("NFD", s) if unicodedata.category(c) != "Mn"
    )


def norm_key(s: str | None) -> str:
    if s is None:
        return ""
    t = strip_accents(str(s).strip()).upper()
    t = re.sub(r"\s+", " ", t)
    return t


def sheet_name_to_code(sheet_name: str) -> str:
    """Aligné sur Kotlin : uppercase puis [^A-Z0-9_] -> _ (max 80)."""
    u = sheet_name.strip().upper()
    code = re.sub(r"[^A-Z0-9_]", "_", u)
    code = re.sub(r"_+", "_", code).strip("_")
    return code[:80] if code else "CORPS"


def load_base_supplier_list(xlsx_path: str) -> list[str]:
    """Noms fournisseurs distincts du BASE (pour résolution des abréviations)."""
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb.active
    suppliers: set[str] = set()
    for row in ws.iter_rows(min_row=2, min_col=5, max_col=8, values_only=True):
        art, u, prix, fourn = row[0], row[1], row[2], row[3]
        if fourn is None and art is None:
            continue
        fa = str(fourn).strip() if fourn is not None else ""
        ar = str(art).strip() if art is not None else ""
        if not ar:
            continue
        if fa:
            suppliers.add(fa)
    wb.close()
    return sorted(suppliers, key=lambda x: x.upper())


def load_base_keys(xlsx_path: str) -> set[tuple[str, str, str]]:
    """Clés (article, unite, fournisseur) du fichier BASE — optionnel pour --skip-if-in-base."""
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb.active
    keys: set[tuple[str, str, str]] = set()
    for row in ws.iter_rows(min_row=2, min_col=5, max_col=8, values_only=True):
        art, u, prix, fourn = row[0], row[1], row[2], row[3]
        if fourn is None and art is None:
            continue
        fa = str(fourn).strip() if fourn is not None else ""
        ar = str(art).strip() if art is not None else ""
        un = str(u).strip() if u is not None else ""
        if not ar:
            continue
        keys.add((norm_key(ar), norm_key(un), norm_key(fa)))
    wb.close()
    return keys


def resolve_supplier(raw: str, canonicals: list[str]) -> str:
    if not raw or not str(raw).strip():
        return raw
    s = str(raw).strip()
    su = strip_accents(s).upper()
    for c in canonicals:
        if strip_accents(c).upper() == su:
            return c
    for c in canonicals:
        cu = strip_accents(c).upper()
        if cu.startswith(su) and len(su) >= 3:
            return c
    for c in canonicals:
        cu = strip_accents(c).upper()
        if len(su) >= 4 and (su in cu):
            return c
    return s


def find_paths():
    base = os.path.join(DOCS, "BASE DE DONNE F.xlsx")
    xls = None
    for f in os.listdir(DOCS):
        if f.endswith(".xls") and "BAREME DES PRIX" in f and "Tout_Corps" in f:
            xls = os.path.join(DOCS, f)
            break
    if not os.path.isfile(base):
        raise FileNotFoundError(base)
    if not xls:
        raise FileNotFoundError("Fichier .xls Tout_Corps dans docs/")
    return base, xls


def iter_material_rows(xls_path: str, sheet_index: int):
    book = xlrd.open_workbook(xls_path, formatting_info=False)
    if sheet_index < 0 or sheet_index >= book.nsheets:
        raise ValueError(f"sheet_index {sheet_index} hors plage (0..{book.nsheets - 1})")
    sh = book.sheet_by_index(sheet_index)
    name = sh.name
    if not name.strip() or name.strip().lower() == "sheet1":
        raise ValueError(f"Feuille index {sheet_index} ({name!r}) ignorée (vide ou Sheet1).")

    def get_cell(r, c):
        if r >= sh.nrows or c >= sh.ncols:
            return None, None
        return sh.cell_value(r, c), sh.cell_type(r, c)

    def cell_str(r, c):
        v, t = get_cell(r, c)
        if t == xlrd.XL_CELL_TEXT:
            return str(v).strip() if v else None
        if t == xlrd.XL_CELL_NUMBER:
            if v == int(v):
                return str(int(v))
            return str(v)
        return None

    def cell_num(r, c):
        v, t = get_cell(r, c)
        if t == xlrd.XL_CELL_NUMBER:
            return float(v)
        return None

    def cell_fourn_text(r, c):
        v, t = get_cell(r, c)
        if t != xlrd.XL_CELL_TEXT:
            return None
        if not v:
            return None
        return str(v).strip() or None

    for ri in range(3, sh.nrows):
        materiaux = cell_str(ri, 2)
        prix = cell_num(ri, 4)
        fourn = cell_fourn_text(ri, 6)
        if not materiaux:
            continue
        if prix is None and not fourn:
            continue
        ref = cell_str(ri, 1) or "—"
        u = cell_str(ri, 3) or "u"
        contact = cell_str(ri, 7) or "—"
        date_prix = cell_str(ri, 5) or None
        prix_bd = Decimal(str(prix)) if prix is not None else Decimal("0")
        yield {
            "row_excel": ri + 1,
            "reference": ref[:50],
            "libelle": materiaux[:2000],
            "unite": u[:20],
            "prix_ttc": prix_bd.quantize(Decimal("0.01")),
            "fournisseur_brut": fourn,
            "contact": contact[:100],
            "date_prix": (date_prix or "")[:50] or None,
            "_sheet_name": name,
        }


def main():
    ap = argparse.ArgumentParser(description="Merge matériaux Tout_Corps par feuille")
    ap.add_argument("--execute", action="store_true", help="Écrire en base")
    ap.add_argument(
        "--sheet-index",
        type=int,
        default=None,
        help="Index xlrd de la feuille (0=coef éloignement, 1=Gros-Oeuvre, 2=Assainissement, …). Si défini, ignore --from/--to.",
    )
    ap.add_argument("--from-index", type=int, default=2, help="Première feuille (défaut 2 = après Gros-Oeuvre)")
    ap.add_argument("--to-index", type=int, default=15, help="Dernière feuille corps d'état")
    ap.add_argument(
        "--skip-if-in-base",
        action="store_true",
        help="Ne pas insérer si la clé (article+unité+fournisseur) existe déjà dans BASE DE DONNE F.xlsx (comportement proche du merge Gros-Œuvre).",
    )
    args = ap.parse_args()
    dry = not args.execute

    base_path, xls_path = find_paths()
    canonical_suppliers = load_base_supplier_list(base_path)
    base_keys = load_base_keys(base_path) if args.skip_if_in_base else None

    book = xlrd.open_workbook(xls_path, formatting_info=False)

    if args.sheet_index is not None:
        indices = [args.sheet_index]
    else:
        indices = list(range(args.from_index, args.to_index + 1))

    # Validation indices
    for idx in indices:
        if idx <= 0:
            print(f"Ignoré : index {idx} (feuille 0 = coef éloignement, non gérée par ce script).", file=sys.stderr)
            sys.exit(1)
        if idx >= book.nsheets:
            print(f"Index {idx} >= nombre de feuilles ({book.nsheets}).", file=sys.stderr)
            sys.exit(1)

    if dry:
        print("Mode dry-run (aucune écriture). Utiliser --execute pour insérer.\n")

    try:
        import pymysql
    except ImportError:
        if not dry:
            print("pymysql requis : pip install pymysql", file=sys.stderr)
            sys.exit(1)

    conn = None
    if not dry:
        env = load_env(ENV_PATH)
        url = env.get("DATABASE_URL", "")
        user = env.get("DATABASE_USERNAME", "root")
        password = env.get("DATABASE_PASSWORD", "")
        if not url:
            print("DATABASE_URL manquant dans backend/.env", file=sys.stderr)
            sys.exit(1)
        host, port, database = parse_jdbc_url(url)
        conn = pymysql.connect(
            host=host,
            port=port,
            user=user,
            password=password,
            database=database,
            charset="utf8mb4",
            autocommit=False,
        )

    total_inserted_lines = 0
    total_inserted_f = 0

    for sheet_index in indices:
        sh = book.sheet_by_index(sheet_index)
        sheet_name = sh.name
        if sheet_name.strip().lower() == "sheet1" and sh.nrows <= 1:
            print(f"[{sheet_index}] {sheet_name!r} : feuille vide, ignorée.")
            continue

        code = sheet_name_to_code(sheet_name)
        to_insert: list[dict] = []
        seen: set[tuple[str, str, str]] = set()

        for row in iter_material_rows(xls_path, sheet_index):
            fb = row["fournisseur_brut"]
            if fb:
                resolved = resolve_supplier(fb, canonical_suppliers)
                if not resolved or resolved.strip() == "":
                    resolved = str(fb).strip()
            else:
                resolved = "Non renseigné"

            k = (norm_key(row["libelle"]), norm_key(row["unite"]), norm_key(resolved))
            if base_keys is not None and k in base_keys:
                continue
            if k in seen:
                continue
            seen.add(k)
            row["fournisseur_resolu"] = resolved
            to_insert.append(row)

        print(f"\n=== Feuille [{sheet_index}] {sheet_name!r} -> code={code!r} ===")
        print(f"   Lignes matière à traiter (hors doublons internes) : {len(to_insert)}")

        if dry:
            for i, r in enumerate(to_insert[:25]):
                print(
                    f"   [{i+1}] {r['libelle'][:55]!r} | {r['unite']!r} | {r['prix_ttc']} | "
                    f"{r['fournisseur_brut']!r} -> {r['fournisseur_resolu']!r}"
                )
            if len(to_insert) > 25:
                print(f"   ... +{len(to_insert) - 25} lignes")
            continue

        assert conn is not None
        try:
            with conn.cursor() as cur:
                cur.execute(
                    "SELECT id FROM bareme_corps_etat WHERE code = %s LIMIT 1",
                    (code,),
                )
                r0 = cur.fetchone()
                if r0:
                    corps_id = r0[0]
                else:
                    cur.execute(
                        "SELECT COALESCE(MAX(ordre_affichage),0)+1 FROM bareme_corps_etat"
                    )
                    ordre_new = cur.fetchone()[0]
                    cur.execute(
                        """
                        INSERT INTO bareme_corps_etat (code, libelle, ordre_affichage, created_at, updated_at)
                        VALUES (%s, %s, %s, NOW(6), NOW(6))
                        """,
                        (code, sheet_name.strip()[:200], ordre_new),
                    )
                    corps_id = cur.lastrowid

                cur.execute(
                    """
                    SELECT LOWER(TRIM(libelle)), LOWER(TRIM(COALESCE(unite,''))),
                           f.id, LOWER(TRIM(COALESCE(f.nom,'')))
                    FROM bareme_lignes_prix l
                    LEFT JOIN bareme_fournisseurs f ON l.fournisseur_bareme_id = f.id
                    WHERE l.corps_etat_id = %s AND l.type = 'MATERIAU' AND l.parent_id IS NULL
                    """,
                    (corps_id,),
                )
                existing = set()
                for lib, un, _fid, fnom in cur.fetchall():
                    existing.add(
                        (norm_key(lib or ""), norm_key(un or ""), norm_key(fnom or ""))
                    )

                cur.execute(
                    "SELECT COALESCE(MAX(ordre_ligne), 0) FROM bareme_lignes_prix WHERE corps_etat_id = %s",
                    (corps_id,),
                )
                ordre = cur.fetchone()[0]
                inserted_this_sheet = 0
                inserted_f_this_sheet = 0

                for r in to_insert:
                    k = (
                        norm_key(r["libelle"]),
                        norm_key(r["unite"]),
                        norm_key(r["fournisseur_resolu"]),
                    )
                    if k in existing:
                        continue
                    nom_f = r["fournisseur_resolu"] or "Non renseigné"
                    cur.execute(
                        "SELECT id FROM bareme_fournisseurs WHERE LOWER(nom) = LOWER(%s) LIMIT 1",
                        (nom_f,),
                    )
                    fr = cur.fetchone()
                    if fr:
                        fid = fr[0]
                    else:
                        cur.execute(
                            """
                            INSERT INTO bareme_fournisseurs (nom, contact, created_at, updated_at)
                            VALUES (%s, %s, NOW(6), NOW(6))
                            """,
                            (nom_f[:200], (r["contact"] or "—")[:100]),
                        )
                        fid = cur.lastrowid
                        total_inserted_f += 1
                        inserted_f_this_sheet += 1

                    ordre += 1
                    dp = r["date_prix"] or ""
                    cur.execute(
                        """
                        INSERT INTO bareme_lignes_prix (
                            corps_etat_id, type, reference, libelle, famille, unite, prix_ttc, date_prix,
                            fournisseur_bareme_id, contact_texte, ordre_ligne, numero_ligne_excel,
                            prix_estime, created_at, updated_at, parent_id
                        ) VALUES (
                            %s, 'MATERIAU', %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, 0, NOW(6), NOW(6), NULL
                        )
                        """,
                        (
                            corps_id,
                            r["reference"],
                            r["libelle"],
                            sheet_name.strip()[:120],
                            r["unite"],
                            str(r["prix_ttc"]),
                            dp[:50] if dp else None,
                            fid,
                            (r["contact"] or "—")[:100],
                            ordre,
                            r["row_excel"],
                        ),
                    )
                    total_inserted_lines += 1
                    inserted_this_sheet += 1
                    existing.add(k)

            conn.commit()
            print(
                f"   OK : +{inserted_this_sheet} ligne(s) matériau, "
                f"+{inserted_f_this_sheet} fournisseur(s) sur cette feuille."
            )
        except Exception as e:
            conn.rollback()
            raise e

    if conn:
        conn.close()

    if dry:
        print("\nDry-run terminé. Relancer avec --execute pour insérer.")
    else:
        print(f"\nTerminé : {total_inserted_f} nouveau(x) fournisseur(s) (cumul), {total_inserted_lines} ligne(s) matériau insérée(s) (cumul).")


if __name__ == "__main__":
    main()
