"""
Import DQE data from Excel files to Render production API.
Usage: python scripts/import_dqe_render.py
"""
import requests
import openpyxl
import time
import sys
import os

API_BASE = "https://mika-services-backend.onrender.com/api"
EMAIL = "olsenkampala@gmail.com"
PASSWORD = "Olsenk2000#2000"

def login():
    r = requests.post(f"{API_BASE}/auth/login", json={"email": EMAIL, "password": PASSWORD}, timeout=30)
    r.raise_for_status()
    return r.json()["accessToken"]

def headers(token):
    return {"Authorization": f"Bearer {token}", "Content-Type": "application/json"}

def create_chapitre(token, projet_id, numero, designation, ordre):
    data = {"projetId": projet_id, "numero": numero, "designation": designation, "ordre": ordre}
    r = requests.post(f"{API_BASE}/dqe/chapitres", json=data, headers=headers(token), timeout=30)
    r.raise_for_status()
    return r.json()["id"]

def create_ligne(token, chapitre_id, numero_poste, designation, unite, quantite, prix_unitaire, montant_total, ordre):
    data = {
        "chapitreId": chapitre_id,
        "numeroPoste": str(numero_poste) if numero_poste else None,
        "designation": designation,
        "unite": unite,
        "quantite": round(quantite, 4) if quantite else 0,
        "prixUnitaire": round(prix_unitaire, 2) if prix_unitaire else 0,
        "montantTotal": round(montant_total, 2) if montant_total else 0,
        "avancementPct": 0,
        "ordre": ordre
    }
    r = requests.post(f"{API_BASE}/dqe/lignes", json=data, headers=headers(token), timeout=30)
    r.raise_for_status()
    return r.json()["id"]

# =============================================
# BEL AIR — Projet ID 2
# =============================================
def import_bel_air(token):
    print("\n=== IMPORT DQE BEL AIR (Projet ID 2) ===")
    projet_id = 2
    wb = openpyxl.load_workbook("docs/DQE BEL AIR.xlsx", data_only=True)
    ws = wb[wb.sheetnames[0]]

    # Chapitre 100 - Travaux Preparatoires
    ch100 = create_chapitre(token, projet_id, "100", "Travaux Preparatoires", 1)
    print(f"  Chapitre 100 cree (id={ch100})")
    lignes_100 = [
        ("101-1", "Installation de chantier - A/R personnel et materiel", "FT", 1, 20000000, 20000000),
        ("101-1b", "Enveloppe QHSE", "FT", 1, 5910000, 5910000),
        ("102-1", "Process documentaire", "FT", 1, 11470000, 11470000),
        ("102-2", "Etudes topo", "FT", 1, 14780000, 14780000),
        ("102-3", "Etudes geotechniques", "FT", 1, 10090000, 10090000),
        ("103-1", "PL eau avec ensemble de compteurs a deplacer", "U", 5, 4170000, 20850000),
        ("103-2", "PL electricite avec ensemble de compteurs a deplacer", "U", 2, 4370000, 8740000),
        ("103-3", "Depose, deplacement de candelabres, toutes sujetions (hors HTA)", "U", 10, 5280000, 52800000),
        ("103-4", "Poteau electrique Haute Tensions (HTA) a deplacer", "U", 2, 6940000, 13880000),
        ("103-5", "Enveloppe des droits de passages", "m2", 100, 30000, 3000000),
    ]
    for i, (num, des, unite, qte, pu, mt) in enumerate(lignes_100, 1):
        create_ligne(token, ch100, num, des, unite, qte, pu, mt, i)
    print(f"  -> {len(lignes_100)} lignes creees (sous-total: 161 520 000)")

    # Chapitre 200 - Terrassements
    ch200 = create_chapitre(token, projet_id, "200", "Terrassements", 2)
    print(f"  Chapitre 200 cree (id={ch200})")
    lignes_200 = [
        ("201", "Depose des paves existants", "m2", 700, 3800, 2660000),
        ("202", "Deblai rocheux entre pk0+750 et pk1+000 epaisseur moyenne de 40cm", "m3", 920, 13820, 12714400),
        ("203", "Peignage des talus a 2/3 a partir du PK0+750 ht moyenne = 3,5m", "m2", 1239, 12260, 15190140),
        ("204", "Mise en forme de PST et compactage", "m2", 11592, 1460, 16924320),
    ]
    for i, (num, des, unite, qte, pu, mt) in enumerate(lignes_200, 1):
        create_ligne(token, ch200, num, des, unite, qte, pu, mt, i)
    print(f"  -> {len(lignes_200)} lignes creees (sous-total: 47 488 860)")

    # Chapitre 300 - Chaussees - Trottoirs
    ch300 = create_chapitre(token, projet_id, "300", "Chaussees - Trottoirs", 3)
    print(f"  Chapitre 300 cree (id={ch300})")
    lignes_300 = [
        ("301", "Rechargement en graveleux lateritique/forme + geotextile", "m3", 1764, 21540, 37996560),
        ("302", "Couche de fondation chaussee en grave non traitee GNT 0/25 ep = 20cm", "m3", 1764, 75240, 132723360),
        ("303", "Couche de base-roulement en beton arme continu type BAC epaisseur 15cm", "m3", 1134, 339590, 385095060),
        ("304", "Trottoir en beton de ciment de 10 cm d'epaisseur 1,4ml cumule", "m2", 1764, 25700, 45334800),
        ("305", "Embranchement de voies (55ml & 20ml & 100ml)", "ml", 175, 472920, 82761000),
    ]
    for i, (num, des, unite, qte, pu, mt) in enumerate(lignes_300, 1):
        create_ligne(token, ch300, num, des, unite, qte, pu, mt, i)
    print(f"  -> {len(lignes_300)} lignes creees (sous-total: 683 910 780)")

    # Chapitre 400 - Assainissement et Drainage
    ch400 = create_chapitre(token, projet_id, "400", "Assainissement et Drainage", 4)
    print(f"  Chapitre 400 cree (id={ch400})")
    lignes_400 = [
        ("401-1", "Creation/prolongement d'exutoires en beton vers marecages (pk0+180)", "ml", 60, 107000, 6420000),
        ("401-2", "Creation/prolongement d'exutoires en caniveaux 80x80 (pk0+375;pk0+600;pk1260)", "ml", 306, 119000, 36414000),
        ("401-3", "Murs de soutenements en beton arme, hauteur moyenne 3,5m (104+115+132+72,6ml)", "ml", 423.6, 509400, 215781840),
        ("401-4", "Travaux sur regards SEEG", "u", 7, 782080, 5474560),
        ("401-5", "Enveloppe des droits de passages", "m2", 500, 30000, 15000000),
        ("402", "Fourniture et pose de bordures de type T2", "ml", 2670, 30000, 80100000),
        ("403", "Fourniture et pose de Caniveaux 50x50", "ml", 1335, 98090, 130950150),
        ("404", "Fourniture et pose de dallettes pour caniveaux 50x50", "ml", 185, 30000, 5550000),
        ("405", "Amenagements divers", "Fft", 1, 148200482, 148200482),
    ]
    for i, (num, des, unite, qte, pu, mt) in enumerate(lignes_400, 1):
        create_ligne(token, ch400, num, des, unite, qte, pu, mt, i)
    print(f"  -> {len(lignes_400)} lignes creees (sous-total: 643 891 032)")

    print("  TOTAL HT BEL AIR: 1 536 810 672 FCFA")
    print("  BEL AIR TERMINE!\n")

# =============================================
# AKEMIDJOGONI — Projet ID 24
# =============================================
def import_akemidjogoni(token):
    print("\n=== IMPORT DQE AKEMIDJOGONI (Projet ID 24) ===")
    projet_id = 24
    wb = openpyxl.load_workbook("docs/DQE FINALISé PROJET AKEMIDJOGONI.xlsx", data_only=True)

    # Lister les feuilles pour comprendre la structure
    print(f"  Feuilles: {wb.sheetnames}")

    # Parcourir chaque feuille comme un chapitre
    for sheet_idx, sheet_name in enumerate(wb.sheetnames, 1):
        ws = wb[sheet_name]
        # Trouver le titre et les lignes de donnees
        titre = sheet_name.strip()

        # Chercher la ligne d'en-tete (N° Prix / Designation / ...)
        header_row = None
        for r in range(1, min(ws.max_row + 1, 20)):
            val = ws.cell(r, 1).value
            if val and str(val).strip().upper() in ("N° PRIX", "N°", "NO", "N\u00b0 PRIX", "N\u00b0"):
                header_row = r
                break

        if header_row is None:
            # Essayer de trouver les donnees directement
            for r in range(1, min(ws.max_row + 1, 20)):
                val = ws.cell(r, 2).value
                if val and str(val).strip().upper() == "DESIGNATION":
                    header_row = r
                    break

        if header_row is None:
            print(f"  [SKIP] Feuille '{titre}' - pas d'en-tete trouve")
            continue

        # Creer le chapitre
        chapitre_num = str(sheet_idx * 100)
        ch_id = create_chapitre(token, projet_id, chapitre_num, titre, sheet_idx)
        print(f"  Chapitre {chapitre_num} '{titre}' cree (id={ch_id})")

        # Lire les lignes de donnees
        ligne_count = 0
        for r in range(header_row + 1, ws.max_row + 1):
            col1 = ws.cell(r, 1).value  # N° Prix
            col2 = ws.cell(r, 2).value  # Designation
            col3 = ws.cell(r, 3).value  # Unite
            col5 = ws.cell(r, 5).value  # Quantite
            col6 = ws.cell(r, 6).value  # Prix unitaire
            col7 = ws.cell(r, 7).value  # Prix total / Montant

            # Sauter les lignes vides ou sous-totaux
            if not col2 or not str(col2).strip():
                continue
            designation = str(col2).strip()
            if designation.upper().startswith(("TOTAL", "SOUS-TOTAL", "TVA", "CSS", "ARRET")):
                continue
            if designation.upper().startswith("CHAPITRE"):
                continue

            # Sauter les lignes de detail (sans N° et sans montant total)
            montant = col7
            if montant is None or not isinstance(montant, (int, float)) or montant == 0:
                # C'est une ligne de detail, on la saute
                continue

            num_poste = str(col1).strip() if col1 else f"L{r}"
            unite = str(col3).strip() if col3 else ""
            quantite = float(col5) if isinstance(col5, (int, float)) else 0
            pu = float(col6) if isinstance(col6, (int, float)) else 0

            ligne_count += 1
            try:
                create_ligne(token, ch_id, num_poste, designation, unite, quantite, pu, float(montant), ligne_count)
            except Exception as e:
                print(f"    [ERR] Ligne {num_poste} '{designation}': {e}")

        print(f"  -> {ligne_count} lignes creees")

    print("  AKEMIDJOGONI TERMINE!\n")


def main():
    os.chdir(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

    print("Connexion a l'API Render...")
    token = login()
    print("Connecte!\n")

    # Import BEL AIR
    try:
        import_bel_air(token)
    except Exception as e:
        print(f"ERREUR BEL AIR: {e}")

    # Re-login au cas ou le token expire
    print("Re-connexion...")
    token = login()

    # Import AKEMIDJOGONI
    try:
        import_akemidjogoni(token)
    except Exception as e:
        print(f"ERREUR AKEMIDJOGONI: {e}")

    print("\n=== IMPORT TERMINE ===")


if __name__ == "__main__":
    main()
