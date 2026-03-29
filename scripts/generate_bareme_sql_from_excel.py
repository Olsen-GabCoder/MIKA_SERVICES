#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Lit BASE_DE_DONNEES_PRO.xlsx (feuille « Base de Données ») et génère un script SQL MySQL
pour peupler bareme_corps_etat, bareme_fournisseurs, bareme_lignes_prix.

Sans table TEMPORARY ni sous-requêtes répétées (compatible MySQL : erreur 1137 évitée).

Usage :
  python scripts/generate_bareme_sql_from_excel.py
  Get-Content backend\\scripts\\sql\\import_bareme_base_de_donnees.sql -Raw -Encoding UTF8 | mysql -u root -p --default-character-set=utf8mb4 mika_services_dev

Prérequis : pip install openpyxl
"""
from __future__ import annotations

from pathlib import Path
from datetime import datetime

try:
    import openpyxl
except ImportError:
    raise SystemExit("Installez openpyxl : pip install openpyxl")

ROOT = Path(__file__).resolve().parents[1]
XLSX = ROOT / "BASE_DE_DONNEES_PRO.xlsx"
OUT = ROOT / "backend" / "scripts" / "sql" / "import_bareme_base_de_donnees.sql"


def normalize_sheet_title(name: str) -> str:
    s = name.strip().lower()
    for a, b in (
        ("é", "e"),
        ("è", "e"),
        ("ê", "e"),
        ("ë", "e"),
        ("à", "a"),
        ("â", "a"),
        ("ù", "u"),
        ("û", "u"),
        ("ô", "o"),
        ("î", "i"),
        ("ç", "c"),
    ):
        s = s.replace(a, b)
    return " ".join(s.split())


def find_base_sheet(wb: openpyxl.Workbook):
    target = normalize_sheet_title("Base de Données")
    for n in wb.sheetnames:
        if normalize_sheet_title(n) == target:
            return wb[n]
    raise SystemExit(f"Feuille « Base de Données » introuvable. Feuilles : {wb.sheetnames}")


def find_header_row(ws) -> int:
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=40, values_only=True)):
        if not row or row[0] is None:
            continue
        a = str(row[0]).strip().upper()
        if "RÉF" in a or a == "REFERENCE" or (a.startswith("REF") and "CORPS" not in str(row)):
            if len(row) > 1 and row[1] and "FOURN" in str(row[1]).upper():
                return i
    raise SystemExit("Ligne d'en-tête (RÉFÉRENCE / FOURNISSEUR) introuvable.")


def sql_str(v) -> str:
    if v is None:
        return "NULL"
    if isinstance(v, bool):
        return "1" if v else "0"
    if isinstance(v, (int, float)):
        if isinstance(v, float) and (v != v):
            return "NULL"
        return str(v)
    s = str(v).replace("\\", "\\\\").replace("'", "''").replace("\x00", "")
    return f"'{s}'"


def to_float(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return None


def main():
    if not XLSX.is_file():
        raise SystemExit(f"Fichier Excel introuvable : {XLSX}")

    wb = openpyxl.load_workbook(XLSX, data_only=True, read_only=True)
    try:
        ws = find_base_sheet(wb)
        header_idx = find_header_row(ws)
        data_rows: list[tuple] = []
        for row in ws.iter_rows(min_row=header_idx + 2, values_only=True):
            if not row or row[0] is None:
                continue
            ref = str(row[0]).strip()
            if not ref.upper().startswith("MAT-"):
                continue
            fourn = str(row[1]).strip() if row[1] else ""
            cat = str(row[2]).strip() if row[2] else ""
            fam = str(row[3]).strip() if row[3] else ""
            des = str(row[4]).strip() if row[4] else ""
            unite = str(row[5]).strip() if row[5] else "u"
            prix = to_float(row[6])
            if not fourn or prix is None:
                continue
            data_rows.append((ref, fourn, cat, fam, des, unite, prix, len(data_rows) + header_idx + 2))
    finally:
        wb.close()

    if not data_rows:
        raise SystemExit("Aucune ligne MAT- valide trouvée.")

    fournisseurs = sorted({r[1] for r in data_rows})
    # IDs explicites fournisseurs : 1..N (tables distinctes de corps_etat id=1)
    fourn_id: dict[str, int] = {nom: i + 1 for i, nom in enumerate(fournisseurs)}

    date_prix = datetime.now().strftime("%Y-%m-%d")
    corps_id = 1

    OUT.parent.mkdir(parents=True, exist_ok=True)
    out_lines: list[str] = [
        "-- Généré par scripts/generate_bareme_sql_from_excel.py",
        "-- Source : BASE_DE_DONNEES_PRO.xlsx, feuille Base de Données",
        f"-- Lignes : {len(data_rows)} | Fournisseurs distincts : {len(fournisseurs)}",
        "-- IDs explicites (corps_etat=1, fournisseurs 1..N) pour éviter erreur MySQL 1137 sur TEMPORARY.",
        "",
        "SET NAMES utf8mb4;",
        "SET FOREIGN_KEY_CHECKS = 0;",
        "",
        "DELETE FROM bareme_lignes_prix;",
        "DELETE FROM bareme_fournisseurs;",
        "DELETE FROM bareme_corps_etat;",
        "",
        "SET FOREIGN_KEY_CHECKS = 1;",
        "",
        "INSERT INTO bareme_corps_etat (id, code, libelle, ordre_affichage, created_at, updated_at)",
        "VALUES (",
        f"  {corps_id},",
        "  'CATALOGUE_BASE_DONNEES',",
        "  'Catalogue Base de données (import Excel PRO)',",
        "  1,",
        "  NOW(6),",
        "  NOW(6)",
        ");",
        "",
    ]

    # Un seul INSERT multi-lignes pour les fournisseurs (IDs fixes)
    fv = []
    for nom in fournisseurs:
        fid = fourn_id[nom]
        fv.append(f"  ({fid}, {sql_str(nom)}, '—', NOW(6), NOW(6))")
    out_lines.append("INSERT INTO bareme_fournisseurs (id, nom, contact, created_at, updated_at) VALUES")
    out_lines.append(",\n".join(fv) + ";")
    out_lines.append("")

    next_auto_f = len(fournisseurs) + 1
    out_lines.append(f"ALTER TABLE bareme_fournisseurs AUTO_INCREMENT = {next_auto_f};")
    out_lines.append("ALTER TABLE bareme_corps_etat AUTO_INCREMENT = 2;")
    out_lines.append("")

    batch_size = 120
    ordre = 0
    batch_vals: list[str] = []

    def flush():
        if not batch_vals:
            return
        out_lines.append(
            "INSERT INTO bareme_lignes_prix (corps_etat_id, type, reference, libelle, unite, prix_ttc, date_prix,"
            " fournisseur_bareme_id, contact_texte, famille, categorie, ref_reception, ordre_ligne, numero_ligne_excel,"
            " prix_estime, created_at, updated_at) VALUES"
        )
        out_lines.append(",\n".join(batch_vals) + ";")
        out_lines.append("")
        batch_vals.clear()

    for ref, fourn, cat, fam, des, unite, prix, num_line_excel in data_rows:
        ordre += 1
        libelle = des if des else "—"
        ref50 = ref[:50]
        cat_sql = sql_str(cat[:120]) if cat else "NULL"
        fam_sql = sql_str(fam[:120]) if fam else "NULL"
        unite20 = (unite[:20] if unite else "u") or "u"
        fid = fourn_id[fourn]
        line = (
            f"  ({corps_id}, 'MATERIAU', {sql_str(ref50)}, {sql_str(libelle)}, {sql_str(unite20)}, "
            f"{prix:.2f}, {sql_str(date_prix)}, {fid}, '—', "
            f"{fam_sql}, {cat_sql}, {sql_str(ref50)}, {ordre}, {num_line_excel}, 0, NOW(6), NOW(6))"
        )
        batch_vals.append(line)
        if len(batch_vals) >= batch_size:
            flush()

    flush()

    out_lines.append("-- Fin import barème")
    out_lines.append("")

    OUT.write_text("\n".join(out_lines), encoding="utf-8")
    print(f"OK : {OUT}")
    print(f"     {len(data_rows)} lignes matériaux, {len(fournisseurs)} fournisseurs.")


if __name__ == "__main__":
    main()
