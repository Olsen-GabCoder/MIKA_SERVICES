"""Analyse rapide BASE_DE_DONNEES_PRO.xlsx pour documentation."""
import openpyxl

path = "BASE_DE_DONNEES_PRO.xlsx"
wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
print("=== FEUILLES ===", wb.sheetnames)
for name in wb.sheetnames:
    ws = wb[name]
    print(f"\n--- Feuille: {name!r} ---")
    rows = list(ws.iter_rows(values_only=True))
    n = len(rows)
    print(f"  lignes: {n}")
    if n == 0:
        continue
    for i, row in enumerate(rows[:20]):
        r = []
        for x in row:
            if x is None:
                r.append("")
            else:
                s = str(x).replace("\n", " ").strip()
                r.append(s[:100] if len(s) > 100 else s)
        while r and r[-1] == "":
            r.pop()
        print(f"  R{i+1}: {r[:30]}{' ...' if len(r) > 30 else ''}")
    maxc = 0
    for row in rows[: min(200, n)]:
        for j, cell in enumerate(row):
            if cell is not None and str(cell).strip() != "":
                maxc = max(maxc, j + 1)
    print(f"  max colonne indexée (échantillon 200 lignes): {maxc}")
wb.close()
